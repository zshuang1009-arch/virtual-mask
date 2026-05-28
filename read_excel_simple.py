
try:
    from openpyxl import load_workbook
    print("openpyxl available!")
    
    wb = load_workbook('数据/社交媒体与体貌焦虑调研_314份样本_真实数据_豆包AI生成.xlsx')
    ws = wb.active
    print(f"Worksheet loaded! Rows: {ws.max_row}, Columns: {ws.max_column}")
    
    print("\nFirst 5 rows:")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            print(row)
    
    print("\nAll column names (first row):")
    first_row = next(ws.iter_rows(values_only=True))
    print(first_row)
    
except ImportError:
    print("openpyxl not available, trying csv fallback...")
    import csv
    print("Need to convert Excel to CSV first")
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()

